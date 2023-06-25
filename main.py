from openpyxl import load_workbook



workbook=load_workbook(filename='temelfinansal.xlsx')



sheet= workbook['isyatirim']


f_col_value = []


a_col_company = []



# F sütunundaki değerleri ve A sütunundaki şirket isimlerini oku


for row in sheet.iter_rows(min_row=2,  values_only=True):


    f_value = row[5]  # F sütunu değeri


    a_value = row[0]  # A sütunu şirket ismi


    if f_value=="PD/DD" or f_value=="A/D":      

        f_col_value.append(100)


    else:

        f_col_value.append(f_value)

    a_col_company.append(a_value)
sorted_values, sirali_sirketler = zip(*sorted(zip(f_col_value, a_col_company)))
for deger, sirket in zip(sorted_values, sirali_sirketler):
    print(deger, sirket)

with open('minimum_pd-dd.txt','w') as file:
    for deger,sirket in zip(sorted_values,sirali_sirketler):


        file.write(f"{sirket},{deger}\n")


